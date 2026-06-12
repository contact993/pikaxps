"""Legacy .xls, Avantage-style sheets, fake-xls text exports, clipboard paste."""
import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook

from xpsfit.io import importer


def make_avantage_sheet(ws):
    """Mimic a Thermo Avantage export: junk rows, name+unit header rows,
    spacer columns, parameter text block on the right."""
    ws["H1"] = "Acquisition Parameters"
    ws["H4"] = "Total acquisition time"
    ws["I4"] = "8 mins 22.5 secs"
    ws["H5"] = "Number of Scans"
    ws["I5"] = 50
    ws["A15"] = "Binding Energy (E)"
    ws["E15"] = "Backgnd."
    ws["A16"] = "eV"
    ws["C16"] = "Counts / s"
    ws["E16"] = "Counts / s"
    be = np.arange(75.08, 55.0, -0.1)
    for i, b in enumerate(be):
        r = 17 + i
        ws.cell(row=r, column=1, value=round(float(b), 2))
        ws.cell(row=r, column=3, value=13000.0 + 10 * i)
        ws.cell(row=r, column=5, value=0.0)
    return be.size


def test_avantage_layout_xlsx(tmp_path):
    p = tmp_path / "avantage.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Ir4f Scan"
    n = make_avantage_sheet(ws)
    wb.save(p)

    assert importer.is_excel_file(p)
    assert importer.excel_engine(p) == "openpyxl"
    df = importer.load_excel(p, "Ir4f Scan")
    assert df.shape == (n, 3)  # spacer/parameter columns dropped
    assert "Binding Energy" in str(df.columns[0])
    assert "Counts" in str(df.columns[1])
    m = importer.guess_mapping(df)
    assert (m.x_col, m.y_col) == (0, 1)
    r = importer.to_region(df, m, name="Ir 4f")
    assert r.x.size == n
    assert np.all(np.diff(r.x) > 0)


def test_real_legacy_xls(tmp_path):
    xlwt = pytest.importorskip("xlwt")
    p = tmp_path / "legacy.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("O1s Scan")
    ws.write(14, 0, "Binding Energy (E)")
    ws.write(15, 0, "eV")
    ws.write(15, 2, "Counts / s")
    for i in range(60):
        ws.write(16 + i, 0, 536.0 - i * 0.1)
        ws.write(16 + i, 2, 5000.0 + i)
    wb.save(str(p))

    assert importer.excel_engine(p) == "xlrd"
    assert importer.excel_sheets(p) == ["O1s Scan"]
    df = importer.load_excel(p, "O1s Scan")
    assert df.shape == (60, 2)
    assert "Binding Energy" in str(df.columns[0])


def test_fake_xls_is_text(tmp_path):
    p = tmp_path / "export.xls"  # instruments love doing this
    p.write_text("BE\tcounts\n284.0\t100\n284.2\t140\n284.4\t120\n", encoding="utf-8")
    assert importer.excel_engine(p) is None
    assert not importer.is_excel_file(p)
    df = importer.load_table(importer.sniff(p))
    assert df.shape == (3, 2)


def test_paste_text_pipeline():
    pasted = "junk header line\nBE\tI\n284.0\t100\n284.1\t120\n284.2\t90\n"
    s = importer.sniff_text(pasted)
    assert s.path is None and s.text == pasted
    assert s.delimiter == "\t"
    assert s.skip_rows == 2
    assert s.header == ["BE", "I"]
    df = importer.load_table(s)
    assert df.shape == (3, 2)
    r = importer.to_region(df, importer.guess_mapping(df), name="paste", source="clipboard")
    assert r.x.size == 3


def test_paste_decimal_numbers_with_commas_fail_gracefully():
    # comma-delimited with no header at all
    s = importer.sniff_text("529.5,1000\n529.6,1100\n")
    assert s.delimiter == ","
    df = importer.load_table(s)
    assert df.shape == (2, 2)
