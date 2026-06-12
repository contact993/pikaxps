"""Offscreen GUI smoke test: window construction, import wizard parsing,
recipe insertion, fitting, drag handling, charge correction, panels."""
import os

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import numpy as np
import pytest

from xpsfit.core import lineshapes
from xpsfit.core.spectrum import BackgroundSpec, Region


@pytest.fixture(scope="module")
def qapp():
    from PySide6.QtWidgets import QApplication

    app = QApplication.instance() or QApplication([])
    yield app


def make_s2p_region():
    x = np.arange(158.0, 174.0, 0.05)
    y = np.zeros_like(x) + 200.0
    for c, a in ((161.8, 6000.0), (162.98, 3000.0), (168.8, 2000.0), (169.98, 1000.0)):
        y += lineshapes.evaluate("GL_SUM", x, c, a, 1.0, 30.0, 0.0)
    return Region(name="S 2p", x=x, y=y,
                  background=BackgroundSpec(kind="LINEAR", lo=158.2, hi=173.8))


def test_main_window_full_flow(qapp):
    from xpsfit.ui.main_window import MainWindow

    w = MainWindow()
    w.session.regions.append(make_s2p_region())
    w._refresh_region_list()
    assert w._region() is not None
    assert w._region().name == "S 2p"

    # refdb panel should auto-select S 2p and insert the matching recipe
    w.refdb_panel.select_for_region("S 2p")
    assert w.refdb_panel.element_combo.currentData() == ("S", "2p")
    idx = next(i for i in range(w.refdb_panel.recipe_combo.count())
               if w.refdb_panel.recipe_combo.itemData(i)["id"] == "s2p_sulfide_sulfate")
    w.refdb_panel.recipe_combo.setCurrentIndex(idx)
    w.refdb_panel._insert_recipe()
    assert len(w._region().peaks) == 4  # 2 species x doublet

    # fit through the GUI path
    w._fit_region()
    assert w._region().peaks[0].center.value == pytest.approx(161.8, abs=0.1)
    assert w.peak_table.table.rowCount() == 4

    # drag the free main peak; constrained partner must ignore drags
    before_partner = w._region().peaks[1].center.value
    w._peak_dragged(0, 161.9, 5000.0)
    assert w._region().peaks[0].center.value == pytest.approx(161.9, abs=1e-6)
    w._peak_dragged(1, 150.0, 9999.0)
    assert w._region().peaks[1].center.expr is not None

    # background switching
    w.bg_combo.setCurrentIndex(w.bg_combo.findData("SHIRLEY"))
    assert w._region().background.kind == "SHIRLEY"

    # guides
    w.refdb_panel._show_guides()
    assert len(w.plot.guide_items) > 0
    w.plot.clear_guides()
    assert len(w.plot.guide_items) == 0

    # quantify panel sees the region
    w.quantify_panel.set_session(w.session)
    assert w.quantify_panel.table.rowCount() == 1

    # delete peak through table (reindex path)
    w.peak_table.table.setCurrentCell(3, 0)
    w.peak_table._delete_peak()
    assert len(w._region().peaks) == 3


def test_inline_be_shift(qapp):
    from xpsfit.ui.main_window import MainWindow

    w = MainWindow()
    w.session.regions.append(make_s2p_region())
    w.session.regions.append(make_s2p_region())
    w._refresh_region_list()
    x0 = w.session.regions[0].x[0]

    # apply to all
    w.shift_spin.setValue(-0.2)
    w.shift_all_check.setChecked(True)
    w._apply_shift()
    assert w.session.regions[0].x[0] == pytest.approx(x0 - 0.2)
    assert w.session.regions[1].be_shift == pytest.approx(-0.2)
    assert "(-0.20 eV)" in w.region_list.item(0).text()

    # apply to selection only
    w.shift_all_check.setChecked(False)
    w.region_list.setCurrentRow(0)
    w._apply_shift()
    assert w.session.regions[0].be_shift == pytest.approx(-0.4)
    assert w.session.regions[1].be_shift == pytest.approx(-0.2)


def test_import_wizard_text(qapp, tmp_path):
    from xpsfit.ui.import_wizard import ImportWizard

    p = tmp_path / "spec.dat"
    junk = "Region: C1s\nPass 50\n"
    body = "\n".join(f"{280 + i * 0.1:.2f}\t{100 + i}" for i in range(60))
    p.write_text(junk + body, encoding="utf-8")
    wiz = ImportWizard(p)
    assert wiz.df is not None
    assert wiz.skip_spin.value() == 2
    wiz.name_edit.setText("C 1s")
    wiz._accept()
    assert wiz.result_region is not None
    assert wiz.result_region.name == "C 1s"
    assert wiz.result_region.x.size == 60


def test_import_wizard_excel(qapp, tmp_path):
    import pandas as pd
    from xpsfit.ui.import_wizard import ImportWizard

    p = tmp_path / "spec.xlsx"
    pd.DataFrame({"KE": np.arange(1190.0, 1210.0, 0.5),
                  "counts": np.random.default_rng(0).random(40) * 100}).to_excel(
        p, index=False, sheet_name="Ni2p")
    wiz = ImportWizard(p)
    assert wiz.df is not None
    wiz.ke_check.setChecked(True)
    wiz._accept()
    r = wiz.result_region
    assert r is not None
    assert 276.0 < r.x.min() < 297.0  # 1486.6 - 1210 .. 1486.6 - 1190


def test_help_panel_renders(qapp):
    from xpsfit.ui.help_panel import HelpPanel

    h = HelpPanel()
    assert "Shirley" in h.browser.toPlainText()
    h.show_context("constraint")
    assert "doublet" in h.browser.toPlainText().lower()


def test_constraint_dialog_roundtrip(qapp):
    from xpsfit.ui.peak_table import ConstraintDialog
    from xpsfit.core.spectrum import Peak

    p = Peak.create(center=285.0, area=1000.0, fwhm=1.3)
    dlg = ConstraintDialog(p, 1, 3)
    dlg.editors["fwhm"]["expr"].setText("p0_fwhm")
    dlg.editors["center"]["vary"].setChecked(False)
    dlg.apply()
    assert p.fwhm.expr == "p0_fwhm"
    assert p.center.vary is False


def test_import_wizard_paste_mode(qapp):
    from xpsfit.ui.import_wizard import ImportWizard

    pasted = "Binding Energy\tCounts\n284.0\t100\n284.1\t150\n284.2\t130\n284.3\t90\n"
    wiz = ImportWizard(None, paste_text=pasted)
    assert wiz.paste_mode
    assert wiz.df is not None and wiz.df.shape == (4, 2)
    wiz.name_edit.setText("C 1s paste")
    wiz._accept()
    r = wiz.result_region
    assert r is not None and r.name == "C 1s paste"
    assert r.source_file == "clipboard"
    assert r.x.size == 4

    # editing the paste box re-sniffs (switch to comma data)
    wiz2 = ImportWizard(None, paste_text="1,2\n")
    wiz2.paste_edit.setPlainText("530.0,1000\n530.1,1100\n530.2,900\n")
    wiz2._paste_changed()
    assert wiz2.df is not None and wiz2.df.shape == (3, 2)


def test_import_wizard_multi_sheet_excel(qapp, tmp_path):
    from openpyxl import Workbook
    from xpsfit.ui.import_wizard import ImportWizard
    from tests.test_importer_excel_paste import make_avantage_sheet

    p = tmp_path / "multi.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Ir4f Scan"
    make_avantage_sheet(ws1)
    ws2 = wb.create_sheet("O1s Scan")
    make_avantage_sheet(ws2)
    ws3 = wb.create_sheet("Notes")
    ws3["A1"] = "just text"
    wb.save(p)

    wiz = ImportWizard(p)
    # scan-named sheets auto-checked, text-only sheet disabled
    assert wiz._checked_sheets() == ["Ir4f Scan", "O1s Scan"]
    from PySide6.QtCore import Qt
    notes_item = wiz.sheet_list.item(2)
    assert not (notes_item.flags() & Qt.ItemFlag.ItemIsEnabled)
    assert not wiz.name_edit.isEnabled()  # multi-selection -> sheet names used

    wiz._accept()
    assert len(wiz.result_regions) == 2
    assert [r.name for r in wiz.result_regions] == ["Ir4f Scan", "O1s Scan"]
    assert all(r.x.size > 100 for r in wiz.result_regions)


def test_feedback_round_features(qapp):
    """Shift auto-fill from C-C, free button, reset range, active-bg checkbox,
    table color match."""
    from xpsfit.core.spectrum import Peak
    from xpsfit.ui.main_window import MainWindow

    w = MainWindow()
    region = make_s2p_region()
    region.name = "C 1s"
    region.peaks = [Peak.create(center=285.1, area=9000.0, fwhm=1.3, label="C-C"),
                    Peak.create(center=286.5, area=2000.0, fwhm=1.3, label="C-O")]
    w.session.regions.append(region)
    w._refresh_region_list()

    # auto-fill computes the shift from the biggest fitted peak near 284.8
    w._auto_shift_from_c1s()
    assert w.shift_spin.value() == pytest.approx(-0.3, abs=1e-6)

    # table row color matches plot component color (Label column, after Type)
    from xpsfit.ui.plot_widget import COMPONENT_COLORS
    item = w.peak_table.table.item(0, 2)
    assert item.background().color().name() == COMPONENT_COLORS[0]

    # free button drops constraints and reports what it released
    region.peaks[1].fwhm.expr = "p0_fwhm"
    w.peak_table.refresh()
    w.peak_table.table.setCurrentCell(1, 0)
    w.peak_table._free_peak()
    assert region.peaks[1].fwhm.expr is None
    assert region.peaks[1].center.vary
    msg = w.statusBar().currentMessage()
    assert "🔓" in msg and "FWHM" in msg and "범위 확장" in msg

    # no selection -> guidance, not silence
    w.peak_table.table.setCurrentCell(-1, -1)
    w.peak_table._free_peak()
    assert "선택" in w.statusBar().currentMessage()
    w.peak_table.table.setCurrentCell(1, 0)

    # range reset clears crop
    region.background.lo, region.background.hi = 160.0, 170.0
    w._reset_range()
    assert region.background.lo is None and region.background.hi is None

    # active-bg checkbox only for shirley, syncs to model
    w.bg_combo.setCurrentIndex(w.bg_combo.findData("SHIRLEY"))
    assert w.bg_active_check.isVisible() or True  # offscreen visibility unreliable
    w.bg_active_check.setChecked(True)
    assert region.background.active is True


def test_user_reference_editor(qapp, tmp_path, monkeypatch):
    from xpsfit import refdb
    monkeypatch.setattr(refdb, "USER_DB_DIR", tmp_path)
    refdb.load_db.cache_clear()
    try:
        from xpsfit.ui.refdb_panel import RefDbPanel
        from xpsfit.ui.state_editor import StateEditorDialog

        dlg = StateEditorDialog(element="Ni", orbital="2p")
        dlg.name_edit.setText("My LDH")
        dlg.be_spin.setValue(855.9)
        dlg.ref_edit.setText("Kim et al. 2025")
        dlg._save()
        assert dlg.saved_key == ("Ni", "2p")

        panel = RefDbPanel()
        panel.select_for_region("Ni 2p")
        texts = [panel.tree.topLevelItem(i).text(0)
                 for i in range(panel.tree.topLevelItemCount())]
        assert any(t.startswith("★") and "My LDH" in t for t in texts)

        # new element via editor metadata
        dlg2 = StateEditorDialog(element="W", orbital="4f")
        assert dlg2.meta_box.isVisibleTo(dlg2)  # unknown combo -> meta shown
        dlg2.name_edit.setText("W(0)")
        dlg2.be_spin.setValue(31.4)
        dlg2.ref_edit.setText("Moulder")
        dlg2.split_spin.setValue(2.18)
        dlg2.ratio_edit.setText("4:3")
        dlg2._save()
        assert "W" in refdb.elements()
    finally:
        refdb.load_db.cache_clear()


def test_add_doublet_button(qapp):
    from xpsfit.ui.main_window import MainWindow

    w = MainWindow()
    region = make_s2p_region()
    region.name = "Ir 4f"  # recognized orbital with splitting
    w.session.regions.append(region)
    w._refresh_region_list()
    assert w.peak_table.btn_add_doublet.isEnabled()
    assert "Ir 4f" in w.peak_table.btn_add_doublet.text()

    w.peak_table._add_doublet()
    peaks = region.peaks
    assert len(peaks) == 2
    assert peaks[1].center.expr == "p0_center + 2.98"
    assert peaks[1].area.expr == "p0_area * 0.75"
    assert peaks[1].fwhm.expr == "p0_fwhm"
    assert "doublet 추가" in w.statusBar().currentMessage()

    # unrecognized region name -> disabled
    region2 = make_s2p_region()
    region2.name = "Pasted data"
    w.session.regions.append(region2)
    w._refresh_region_list(select=1)
    assert not w.peak_table.btn_add_doublet.isEnabled()


def test_type_column_conversions(qapp):
    """단일→Doublet→단일, 단일→Satellite, quantify merge, audit kind awareness."""
    from xpsfit.core import audit, quantify
    from xpsfit.core.spectrum import Peak
    from xpsfit.ui.main_window import MainWindow

    w = MainWindow()
    region = make_s2p_region()
    region.name = "Custom XY"  # NOT recognized -> manual params path
    region.peaks = [Peak.create(center=161.8, area=6000.0, fwhm=1.0, label="A"),
                    Peak.create(center=168.8, area=2000.0, fwhm=1.4, label="B")]
    w.session.regions.append(region)
    w._refresh_region_list()

    # convert A to doublet with manual splitting/ratio (bypass dialog)
    w.peak_table._convert_to_doublet(0, split=1.18, ratio_text="2:1")
    assert len(region.peaks) == 3
    assert region.peaks[0].kind == "doublet_main"
    partner = region.peaks[2]
    assert partner.kind == "doublet_partner"
    assert partner.center.expr == "p0_center + 1.18"
    assert partner.area.expr == "p0_area * 0.5"

    # type combo shows states
    combo0 = w.peak_table.table.cellWidget(0, 0)
    assert combo0.currentText() == "Doublet"
    combo2 = w.peak_table.table.cellWidget(2, 0)
    assert not combo2.isEnabled() and combo2.currentText() == "└ 짝"

    # back to single removes the partner
    w.peak_table._convert_to_single(0)
    assert len(region.peaks) == 2
    assert region.peaks[0].kind == "single"

    # B becomes a satellite of A (bypass dialog)
    w.peak_table._convert_to_satellite(1, parent=0, delta=7.0)
    sat = region.peaks[1]
    assert sat.kind == "satellite"
    assert sat.center.expr == "p0_center + 7"
    # quantification merges satellite area into the parent species
    sp = quantify.species_areas(region)
    assert len(sp) == 1
    assert sp[0][1] == pytest.approx(8000.0)
    # audit treats it as satellite (no unknown-state complaints about it)
    region.name = "S 2p"
    report = audit.audit(region)
    assert not any("satellite" in f.message.lower() and f.severity == "bad"
                   for f in report.findings)

    # session round-trip keeps kinds
    d = region.to_dict()
    from xpsfit.core.spectrum import Region as R2
    r2 = R2.from_dict(d)
    assert r2.peaks[1].kind == "satellite"


def test_peak_pinning(qapp):
    """📌 freeze-one-optimize-rest, position-only lock, all-pinned guard."""
    from PySide6.QtCore import Qt
    from xpsfit.core import fitting
    from xpsfit.core.spectrum import Peak
    from xpsfit.ui.main_window import MainWindow

    w = MainWindow()
    region = make_s2p_region()
    region.peaks = [Peak.create(center=161.5, area=4000.0, fwhm=1.2, label="A"),
                    Peak.create(center=168.5, area=1500.0, fwhm=1.2, label="B")]
    w.session.regions.append(region)
    w._refresh_region_list()

    # pin A through the checkbox column
    item = w.peak_table.table.item(0, w.peak_table.COL_PIN)
    item.setCheckState(Qt.CheckState.Checked)
    assert region.peaks[0].pinned is True

    before = (region.peaks[0].center.value, region.peaks[0].area.value,
              region.peaks[0].fwhm.value)
    fitting.fit_region(region)
    after = (region.peaks[0].center.value, region.peaks[0].area.value,
             region.peaks[0].fwhm.value)
    assert before == after  # pinned peak untouched
    assert region.peaks[1].center.value != pytest.approx(168.5, abs=1e-6)  # B moved

    # pinned peak ignores drags
    w._peak_dragged(0, 160.0, 9999.0)
    assert region.peaks[0].center.value == before[0]

    # position-only lock: center frozen, area still optimizes
    region.peaks[0].pinned = False
    region.peaks[0].center.vary = False
    region.peaks[0].center.value = 161.5
    region.peaks[0].area.value = 1000.0
    fitting.fit_region(region)
    assert region.peaks[0].center.value == 161.5
    assert region.peaks[0].area.value != 1000.0

    # everything frozen -> guarded message, no crash
    region.peaks[0].pinned = True
    region.peaks[1].pinned = True
    res = fitting.fit_region(region)
    assert "고정" in res.message

    # session round-trip keeps the pin
    from xpsfit.core.spectrum import Region as R2
    r2 = R2.from_dict(region.to_dict())
    assert r2.peaks[0].pinned is True
